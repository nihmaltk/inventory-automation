import openpyxl
from src.config import *

# Function for loading excel workbook from the specified file path.
def load_workbook(file_path):

    try:
        workbook = openpyxl.load_workbook(file_path)
        return workbook
    except FileNotFoundError:
        print(f"Error: File '{file_path}' not found.")
        raise
    except Exception as e:
        print(f"Error loading workbook: {e}")
        raise

# Function to get a specific sheet from the workbook.
def get_sheet(workbook, sheet_name):

    try:
        return workbook[sheet_name]
    except KeyError:
        print(f"Error: Sheet '{sheet_name}' not found in workbook.")
        raise

# Function to save the workbook to the specified file path.
def save_workbook(workbook, file_path):

    try:
        workbook.save(file_path)
        print(f"File saved successfully to: {file_path}")
    except Exception as e:
        print(f"Error saving workbook: {e}")
        raise

# Function to calculate product counts per supplier.
def calculate_product_counts(sheet):

    product_count_per_supplier = {}
    
    for row in range(2, sheet.max_row + 1):
        supplier_name = sheet.cell(row, COL_SUPPLIER_NAME).value
        
        if supplier_name in product_count_per_supplier:
            product_count_per_supplier[supplier_name] += 1
        else:
            product_count_per_supplier[supplier_name] = 1
    
    return product_count_per_supplier

# Function to calculate total inventory values per supplier.
def calculate_inventory_values(sheet):
    
    total_inventory_value_per_supplier = {}
    
    for row in range(2, sheet.max_row + 1):
        supplier_name = sheet.cell(row, COL_SUPPLIER_NAME).value
        inventory = sheet.cell(row, COL_INVENTORY).value
        price = sheet.cell(row, COL_PRICE).value
        
        inventory_value = inventory * price
        
        if supplier_name in total_inventory_value_per_supplier:
            total_inventory_value_per_supplier[supplier_name] += inventory_value
        else:
            total_inventory_value_per_supplier[supplier_name] = inventory_value
    
    return total_inventory_value_per_supplier

# Function to detect low inventory products.
def get_low_inventory_products(sheet):
    
    low_inventory_products_list = []
    
    for row in range(2, sheet.max_row + 1):
        inventory = sheet.cell(row, COL_INVENTORY).value
        reorder_level = sheet.cell(row, COL_REORDER_LEVEL).value
        
        if inventory < reorder_level:
            product_id = sheet.cell(row, COL_PRODUCT_ID).value
            product_name = sheet.cell(row, COL_PRODUCT_NAME).value
            supplier_name = sheet.cell(row, COL_SUPPLIER_NAME).value
            
            low_inventory_product = {
                'product_id': product_id,
                'product_name': product_name,
                'supplier_name': supplier_name,
                'inventory': inventory,
                'reorder_level': reorder_level,
                'shortage': reorder_level - inventory
            }
            low_inventory_products_list.append(low_inventory_product)
    
    return low_inventory_products_list

# Function to add computed columns to the sheet.
def add_computed_columns(sheet):

    # Add headers
    sheet.cell(1, COL_RESTOCK).value = HEADER_RESTOCK
    sheet.cell(1, COL_INVENTORY_VALUE).value = HEADER_INVENTORY_VALUE
    
    # Calculate and add values
    for row in range(2, sheet.max_row + 1):
        inventory = sheet.cell(row, COL_INVENTORY).value
        price = sheet.cell(row, COL_PRICE).value
        reorder_level = sheet.cell(row, COL_REORDER_LEVEL).value
        
        # Restock status
        restock = 'Yes' if inventory < reorder_level else 'No'
        sheet.cell(row, COL_RESTOCK).value = restock
        
        # Inventory value
        inventory_value = inventory * price
        sheet.cell(row, COL_INVENTORY_VALUE).value = inventory_value

# Function to create a summary sheet.
def create_summary_sheet(workbook, sheet):
    
    # Calculate statistics
    product_counts = calculate_product_counts(sheet)
    inventory_values = calculate_inventory_values(sheet)
    
    # Create summary sheet
    summary_sheet = workbook.create_sheet(SUMMARY_SHEET_NAME)
    
    # Add headers
    for col, header in enumerate(SUMMARY_HEADERS, start=1):
        summary_sheet.cell(1, col).value = header
    
    # Add data
    row_index = 2
    for supplier in product_counts:
        summary_sheet.cell(row_index, 1).value = supplier
        summary_sheet.cell(row_index, 2).value = product_counts[supplier]
        summary_sheet.cell(row_index, 3).value = inventory_values[supplier]
        row_index += 1